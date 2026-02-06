from typing import Any

from pathlib import Path
from openpyxl import Workbook


class ExcelExportService:
    @staticmethod
    def write_amount_statistics(amount_dict: dict, output_path: str | Path = "report.xlsx") -> Path:
        """
        amount_dict: {amount_str: {"count": int, "amount_value": str, "items": [...]}}
        output_path: file path (template, "report.xlsx")

        returns: path to the file
        """
        title = "Amount Statistics | SmartKasa"

        transformed_data = [{"count": amount_element["count"], "amount_value" : amount_element["amount_value"]} for amount_key, amount_element in amount_dict.items() ]

        return ExcelExportService.write_sheet(data=transformed_data, output_path=output_path, sheet_name=title)
    
    @staticmethod
    def write_sheet(data: list[dict], output_path: str | Path, sheet_name: str = "Sheet1", headers_dict: dict[str, str] | None = None) -> Path:
        """
        Writes data to new sheet in the workbook with headers as keys
        
        return: path to the file
        """
        path = Path(output_path)
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if data and isinstance(data, list) and len(data) > 0:
            if headers_dict:
                headers = list(headers_dict.keys())
                headers_keys = list(headers_dict.values())
            else:
                headers = list(data[0].keys())
                headers_keys = headers
            
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)

            for row_idx, row_data in enumerate(data, start=2):
                for col_idx, header_key in enumerate(headers_keys, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(header_key))

        wb.save(path)
        return path